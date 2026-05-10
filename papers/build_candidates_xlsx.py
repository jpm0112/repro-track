"""
Generate `candidates.xlsx` from `candidates.yaml`.

Reads the curated candidate list, HTTP-verifies all link fields
(doi, arxiv_id, openreview_url, code_url, project_page), stamps
`verified_on` with today's ISO date when every non-empty URL returns
a 2xx/3xx status, and writes a sorted Excel sheet (NeurIPS first,
then by fit score) with a frozen header and autofilter.

Usage
-----
    python papers/build_candidates_xlsx.py

Outputs
-------
    papers/candidates.xlsx

Inputs
------
    papers/candidates.yaml (hand-curated, source of truth)

Dependencies
------------
    pyyaml, requests, openpyxl
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path
from typing import Any

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------- configuration ----------

# Column order matches the plan in C:\Users\JP\.claude\plans\adaptive-discovering-pond.md
COLUMNS: list[str] = [
    "id",
    "title",
    "authors",
    "venue",
    "main_track",
    "year",
    "doi",
    "arxiv_id",
    "openreview_url",
    "code_url",
    "code_license",
    "project_page",
    "tldr",
    "method_summary",
    "central_claim",
    "headline_metric",
    "datasets",
    "backbone_models",
    "compute_reported",
    "episodic_to_semantic_fit",
    "repro_difficulty",
    "mlrc_eligibility",
    "risks",
    "notes",
    "verified_on",
]

# venue rank: lower = higher priority (drives row sort order)
# user preference: NeurIPS first, then ICLR/ICML, then ACL-family, then IJCAI/AAAI, then workshops, then arXiv
# Matched substring (case-insensitive); first match wins so order more-specific patterns first.
VENUE_RANK_RULES: list[tuple[str, int]] = [
    # NeurIPS main (presentation tier matters)
    ("neurips 2025 (main, oral",       1),
    ("neurips 2025 (main, spotlight",  2),
    ("neurips 2025 (main, poster",     3),
    ("neurips 2025 (main",             3),
    # ICLR main
    ("iclr 2025 (main, oral",          4),
    ("iclr 2025 (main, spotlight",     5),
    ("iclr 2025 (main",                6),
    ("iclr 2025",                      6),
    # ICML main
    ("icml 2025 (main, oral",          7),
    ("icml 2025",                      8),
    # ACL family main (ACL > EMNLP > NAACL by convention)
    # IMPORTANT: NAACL contains "acl" as substring, so list NAACL rules BEFORE ACL rules.
    ("naacl 2025 (main",              11),
    ("naacl 2025",                    11),
    ("emnlp 2025 (findings",          13),
    ("emnlp 2025 (main, oral",        10),
    ("emnlp 2025 (main",              10),
    ("emnlp 2025",                    10),
    ("acl 2025 (findings",            12),
    ("acl 2025 (main, long",           9),
    ("acl 2025 (main",                 9),
    ("acl 2025",                       9),
    # Other top conferences
    ("ijcai 2025",                    14),
    ("aaai 2025",                     15),
    # Journals
    ("tmlr 2025",                     16),
    ("jmlr 2025",                     17),
    # NeurIPS workshops
    ("neurips 2025 workshop",         30),
    # ICLR 2026 (pending acceptance, future venue)
    ("iclr 2026",                     50),
    # arXiv (last)
    ("arxiv",                         99),
]

REQUEST_TIMEOUT_S = 15
USER_AGENT = "repro-track-link-verifier/0.1 (+https://github.com/jpm0112)"


# ---------- helpers ----------

def venue_rank(venue: str) -> int:
    """Return the sort rank for a venue string. Lower is higher priority.

    Uses case-insensitive substring matching against VENUE_RANK_RULES;
    first matching rule wins, so list more-specific patterns first.
    """
    v = venue.lower()
    for needle, rank in VENUE_RANK_RULES:
        if needle in v:
            return rank
    return 60  # unknown venue, between known ones and arXiv


def doi_to_url(doi: str) -> str:
    """Convert a bare DOI to a doi.org URL. Pass-through if already a URL."""
    doi = doi.strip()
    if not doi:
        return ""
    if doi.startswith("http"):
        return doi
    return f"https://doi.org/{doi}"


def arxiv_to_url(arxiv_id: str) -> str:
    """Convert an arXiv ID (e.g. 2407.09450) to its abs page URL."""
    arxiv_id = arxiv_id.strip()
    if not arxiv_id:
        return ""
    return f"https://arxiv.org/abs/{arxiv_id}"


def verify_url(url: str, session: requests.Session) -> tuple[bool, int | str]:
    """
    HTTP-check a URL. Returns (ok, status_or_error).

    Some sites refuse HEAD; fall back to a streamed GET in those cases.
    """
    if not url:
        return True, "empty"  # nothing to verify
    try:
        resp = session.head(url, allow_redirects=True, timeout=REQUEST_TIMEOUT_S)
        # some servers (e.g. doi.org via CrossRef) return 405/403 on HEAD
        if resp.status_code in (200, 301, 302, 303, 307, 308):
            return True, resp.status_code
        # try GET as fallback
        resp = session.get(url, allow_redirects=True, timeout=REQUEST_TIMEOUT_S, stream=True)
        resp.close()
        return resp.status_code < 400, resp.status_code
    except requests.RequestException as exc:
        return False, f"{type(exc).__name__}: {exc}"


def verify_candidate(cand: dict[str, Any], session: requests.Session) -> dict[str, tuple[bool, Any]]:
    """Verify all link fields for one candidate. Returns dict of field -> (ok, info)."""
    checks = {
        "doi": doi_to_url(cand.get("doi", "")),
        "arxiv_id": arxiv_to_url(cand.get("arxiv_id", "")),
        "openreview_url": cand.get("openreview_url", "").strip(),
        "code_url": cand.get("code_url", "").strip(),
        "project_page": cand.get("project_page", "").strip(),
    }
    results = {}
    for field, url in checks.items():
        ok, info = verify_url(url, session)
        results[field] = (ok, info)
    return results


def autosize_columns(ws: Worksheet, max_width: int = 60) -> None:
    """Approximate auto-sizing: width = min(longest cell, max_width)."""
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        longest = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
            v = row[0]
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = min(longest + 2, max_width)


def style_header(ws: Worksheet) -> None:
    """Bold + colored header row, frozen + autofilter."""
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def wrap_data_cells(ws: Worksheet) -> None:
    """Wrap text in all data cells so long fields are readable."""
    align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = align


# ---------- main ----------

def main() -> int:
    papers_dir = Path(__file__).resolve().parent
    yaml_path = papers_dir / "candidates.yaml"
    xlsx_path = papers_dir / "candidates.xlsx"

    if not yaml_path.exists():
        print(f"ERROR: {yaml_path} not found", file=sys.stderr)
        return 2

    with yaml_path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    candidates: list[dict[str, Any]] = data.get("candidates", [])
    if not candidates:
        print("ERROR: no candidates found in YAML", file=sys.stderr)
        return 2

    today_iso = dt.date.today().isoformat()

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    print(f"Verifying links for {len(candidates)} candidates...")
    all_failures: list[str] = []
    for cand in candidates:
        results = verify_candidate(cand, session)
        failed = [f for f, (ok, _info) in results.items() if not ok]
        if failed:
            for f in failed:
                ok, info = results[f]
                msg = f"  [FAIL] {cand['id']:25} {f:15} {info}"
                print(msg)
                all_failures.append(msg)
            cand["verified_on"] = ""
        else:
            cand["verified_on"] = today_iso
            print(f"  [OK]   {cand['id']:25} all links verified")

    # sort: main-track first, then by venue rank, then descending fit, then ascending repro difficulty
    candidates.sort(key=lambda c: (
        0 if c.get("main_track", False) else 1,
        venue_rank(c.get("venue", "")),
        -int(c.get("episodic_to_semantic_fit", 0) or 0),
        int(c.get("repro_difficulty", 5) or 5),
    ))

    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"

    ws.append(COLUMNS)
    for cand in candidates:
        ws.append([cand.get(col, "") for col in COLUMNS])

    style_header(ws)
    wrap_data_cells(ws)
    autosize_columns(ws)

    wb.save(xlsx_path)
    print(f"\nWrote {xlsx_path}  ({len(candidates)} rows)")
    if all_failures:
        print(f"\n{len(all_failures)} broken link(s) - see [FAIL] lines above")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
